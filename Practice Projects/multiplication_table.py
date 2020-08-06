#! /usr/bin/env python3
# multiplication_table.py - Takes a number N from the cli and creates
# an N*N multiplication table in an Excel spreadsheet.

from openpyxl.styles import Font
import openpyxl
import sys


def create_table():
    # Create Workbook and Worksheet objects, assign a variable to N, give sheet a title
    wb = openpyxl.Workbook()
    sheet = wb.active
    n = int(sys.argv[1])
    sheet.title = f"Multiplcation Table for {n}"
    bold_font = Font(bold=True)  # Make bold font type object
    # Make row/col setup for the N*N table
    for i in range(1, n + 1):
        sheet[f"A{i+1}"].font = bold_font
        sheet[f"A{i+1}"] = i
        sheet.cell(row=1, column=i + 1).font = bold_font
        sheet.cell(row=1, column=i + 1).value = i
    # Insert function in each cell to do the math
    for row in range(1, n + 1):
        for col in range(1, n + 1):
            _ = sheet.cell(row=row + 1, column=col + 1, value=f"={row}*{col}")
    wb.save(f"Multiplication_table{n}.xlsx")


def main():
    create_table()


if __name__ == "__main__":
    main()
