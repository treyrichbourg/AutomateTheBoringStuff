#! /usr/bin/env python3
# spreadsheet_cell_inverter.py - Inverts the row and column of the cells in a spreadsheet

import sys
import openpyxl
from pathlib import Path

# Get args
def get_filename():
    if len(sys.argv) == 2:
        filename = sys.argv[1]
        filename = Path(filename).absolute()
        return filename


# Invert Cells
def cell_inverter(filename):
    # Get workbook/sheet
    wb = openpyxl.load_workbook(filename)
    active_sheet = wb.active
    # Get max rows and columns
    max_rows = active_sheet.max_row
    max_cols = active_sheet.max_column
    # Make new sheet
    invert_sheet = wb.create_sheet("Inverted")
    # Loop through cells in active sheet and update new sheet
    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            invert_sheet.cell(row=col, column=row).value = active_sheet.cell(
                row=row, column=col
            ).value
    wb.save(filename)


def main():
    filename = get_filename()
    cell_inverter(filename)


if __name__ == "__main__":
    main()

