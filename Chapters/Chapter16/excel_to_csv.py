#! /usr/bin/env python3
# excel_to_csv.py - Reads all the Excel files in the current working directory
# and outputs them as CSV files

import openpyxl
import csv
import os

for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):  continue
    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        #Loop through every sheet in the workbook.
        sheet = wb[sheet_name]

        # Create the CSV filename from the Excel filename and sheet title.
        excel_filename = excel_file.replace('.xlsx', '') 
        csv_file = open(f'{excel_filename}_{sheet_name}.csv', 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csv_writer = csv.writer(csv_file)
        # Loop through every row in the sheet.
        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            # Loop through each cell in the row.
            for col_num in range(1, sheet.max_column + 1):
                # Append each cell's data to row_data.
                row_data.append(sheet.cell(row=row_num, column=col_num).value)
            # Write the row_data to the CSV file.
            csv_writer.writerow(row_data)
        csv_file.close()